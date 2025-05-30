from PySide6.QtWidgets import QApplication, QMainWindow, QFileDialog, QGraphicsPixmapItem, QGraphicsScene, QMessageBox,  QVBoxLayout,  QSizePolicy
from PySide6.QtGui import QPixmap, QImage, QIcon, QWindow
import time
from PySide6.QtCore import Qt, QSize
from ui_form import Ui_MainWindow
from customGraphicsView import ImageGraphicsView
import sys
import cv2
import numpy as np
import imagej
import scyjava as sj
import os


from PIL import Image
import pillow_heif



# from jnius import autoclass
import scyjava


import shutil
import hashlib
from PySide6.QtWidgets import  QListWidget, QListWidgetItem, QWidget

import win32gui

import win32con
import pandas as pd


import tkinter as tk
from tkinter import filedialog

import json
from openpyxl import Workbook, load_workbook



pillow_heif.register_heif_opener()




class ImageUploader(QMainWindow):
    def __init__(self, app_folder="app_images"):
        super().__init__()
        self.app_folder = app_folder
        self.saved_hashes = {}  # Dictionary to store hashes of saved images with the corresponding file path
        self.metadata_path = "metadata/saved_hashes.json"


        # Load existing hashes from the folder if any
        # self._load_existing_hashes()
        self.load_metadata()

        if not os.path.exists(self.app_folder):
            os.makedirs(self.app_folder)

        # Set up the main window layout
        self.setWindowTitle("Image Uploader")
        self.setGeometry(100, 100, 800, 600)

        # List Widget to show images
        self.image_list_widget = QListWidget(self)
        self.image_list_widget.setGeometry(10, 10, 780, 550)
        self.image_list_widget.setViewMode(QListWidget.IconMode)
        self.image_list_widget.setIconSize(QSize(100, 100))
        self.image_list_widget.setSpacing(10)
        self.image_list_widget.setDragEnabled(False)  # Disable dragging on the list widget
        self.image_list_widget.setAcceptDrops(False)  # Also ensure drops are not accepted

        # Load images into the list widget
        self.load_images_into_widget()

        # Set layout and central widget
        layout = QVBoxLayout()
        layout.addWidget(self.image_list_widget)
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)
    def load_metadata(self):
            """Load saved_hashes from metadata JSON file."""
            if os.path.exists(self.metadata_path):
                try:
                    with open(self.metadata_path, "r") as f:
                        self.saved_hashes = json.load(f)
                except Exception as e:
                    print(f"Failed to load saved_hashes: {e}")
            else:
                print("No saved_hashes metadata found, starting fresh.")
    def get_all_images(self):
        return [meta["path"] for meta in self.saved_hashes.values()]

    def _load_existing_hashes(self):
        """Load the hashes of images already in the app folder"""
        for image_name in os.listdir(self.app_folder):
            image_path = os.path.join(self.app_folder, image_name)
            if os.path.isfile(image_path):
                image_hash = self._get_image_hash(image_path)
                self.saved_hashes[image_hash] = {
                    "path": image_path,
                    "project_id": None,
                    "colony_label": None
                }
    #Hash each uploaded image, so that duplicates cannot be uploaded
    def _get_image_hash(self, image_path):
        """Generate a hash for the given image file"""
        hash_sha256 = hashlib.sha256()
        with open(image_path, 'rb') as f:
            while chunk := f.read(8192):  # Read file in chunks
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()

    def upload_images(self):
        """Upload and save images, skipping duplicates based on hash"""
        # Open a file dialog to select images
        file_paths, _ = QFileDialog.getOpenFileNames(None, "Select Images", "", "Images (*.png *.jpg *.jpeg *.bmp *.tiff *.heic)")
        if not file_paths:
            return

        saved_images = []

        for image_path in file_paths:
            try:
                # Generate the hash of the current image
                image_hash = self._get_image_hash(image_path)

                # Check if the image hash already exists
                if image_hash in self.saved_hashes:
                    existing_image = self.saved_hashes[image_hash]["path"]  # Get the existing image path
                    print(f"Image {os.path.basename(image_path)} is already saved as {os.path.basename(existing_image)}.")
                    continue  # Skip this image as it is a duplicate based on content

                # Proceed to save the image
                image_name = os.path.basename(image_path)
                saved_path = os.path.join(self.app_folder, image_name)

                # Ensure unique name if image already exists (based on name)
                if os.path.exists(saved_path):
                    base_name, ext = os.path.splitext(image_name)
                    counter = 1
                    while os.path.exists(saved_path):
                        new_name = f"{base_name}_{counter}{ext}"
                        saved_path = os.path.join(self.app_folder, new_name)
                        counter += 1

                shutil.copy(image_path, saved_path)
                self.saved_hashes[image_hash] = {
                    "path": saved_path,
                    "project_id": None,
                    "colony_label": None
                }
                saved_images.append(saved_path)

            except Exception as e:
                QMessageBox.warning(None, "Error", f"Failed to save image {image_path}: {str(e)}")

        # Save the updated saved_hashes dict to the JSON file in metadata folder
        if saved_images:
            import json
            metadata_dir = "metadata"
            os.makedirs(metadata_dir, exist_ok=True)
            metadata_path = os.path.join(metadata_dir, "saved_hashes.json")

            try:
                with open(metadata_path, "w") as f:
                    json.dump(self.saved_hashes, f, indent=4)
                print(f"Saved hashes metadata to {metadata_path}")
            except Exception as e:
                QMessageBox.warning(None, "Error", f"Failed to save metadata JSON file: {str(e)}")

            QMessageBox.information(None, "Images Saved", f"Images have been successfully saved to: {self.app_folder}")

        return saved_images  # List of saved image paths



    def load_images_into_widget(self):
        """Load the images from the folder into the list widget"""
        for image_hash, meta in self.saved_hashes.items():
            image_path = meta["path"]
            # Create a list item for each image
            image_name = os.path.basename(image_path)
            item = QListWidgetItem(image_name)

            # Create a thumbnail for each image
            print(f"image_path type: {type(image_path)}, value: {image_path}")
            pixmap = QPixmap(image_path)
            if pixmap.isNull():
                print(f"Failed to load image: {image_path}")
                continue  # Skip this image if it can't be loaded
            pixmap = pixmap.scaled(100, 100, Qt.KeepAspectRatio)  # Scale the pixmap to fit in the list widget
            item.setIcon(QIcon(pixmap))

            # Add item to the list widget
            self.image_list_widget.addItem(item)

            # Make the item clickable and associate the image path with it
            item.setData(Qt.UserRole, image_path)  # Store the path of the image in the item data
            print(f"Image added to list: {image_name}")  # Debugging print statement

        # Connect the itemClicked signal to the handler after adding the items
        self.image_list_widget.itemClicked.connect(self.on_image_clicked)



    def on_image_clicked(self, item):
        """Handle the click on an image item in the list"""
        image_path = item.data(Qt.UserRole)  # Retrieve the image path from the item data
        print(f"Clicked on image: {image_path}")  # This should print when you click an image item

        # Call the function to process the clicked image
        self.process_single_image(image_path)

    def process_single_image(self, image_path):
        """Process the selected image based on the path"""
        print(f"Processing {image_path}")  # Debug: Check which image is being processed

        # Load the image using Pillow (with pillow-heif support)
        image = Image.open(image_path)

        # Convert the image to QPixmap to display in the QGraphicsView
        image = image.convert("RGB")  # Ensure it's in RGB mode
        data = image.tobytes("raw", "RGB")  # Convert to raw byte data
        qim = QImage(data, image.width, image.height, image.width * 3, QImage.Format_RGB888)

        # Convert QImage to QPixmap
        pixmap = QPixmap.fromImage(qim)

        if pixmap.isNull():
            print("Failed to load image!")  # Debug: Image not loaded properly
            return

        # Clear any existing items in the scene
        self.scene.clear()

        # Create a QGraphicsPixmapItem with the selected image and add it to the scene
        self.image_item = QGraphicsPixmapItem(pixmap)
        self.scene.addItem(self.image_item)

        # Fit the image to the QGraphicsView initially
        self.image_container.fitInView(self.image_item, Qt.KeepAspectRatio)

        # Store the pixmap for potential future use
        self.pixmap = pixmap

class MainWindow(QMainWindow):

    def display_thumbnails(self, image_paths):
        """Displays thumbnails of the images in the QListWidget."""
        self.ui.thumbnailWidget.clear()
        for path in image_paths:
            item = QListWidgetItem()

            # Load with QPixmap first; if fails (e.g. HEIC), fallback to Pillow
            pixmap = QPixmap(path)
            if pixmap.isNull():
                # Try loading with Pillow if QPixmap fails (common with HEIC)
                try:
                    from PIL import Image
                    pil_img = Image.open(path)
                    pil_img.thumbnail((100, 100))
                    pil_img = pil_img.convert("RGBA")
                    data = pil_img.tobytes("raw", "RGBA")
                    qimage = QImage(data, pil_img.width, pil_img.height, QImage.Format.Format_RGBA8888)
                    pixmap = QPixmap.fromImage(qimage)
                except Exception as e:
                    print(f"Failed to load image {path}: {e}")
                    continue

            pixmap = pixmap.scaled(100, 100, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            icon = QIcon(pixmap)
            item.setIcon(icon)
            item.setText(os.path.basename(path))
            item.setSizeHint(QSize(120, 120))  # Add some padding
            item.setData(Qt.UserRole, path)  # Store the image path as custom data
            self.ui.thumbnailWidget.addItem(item)



    def __init__(self):
        super().__init__()
        self.hwnd = []
        sj.config.add_option('-Xmx1g')
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ij = imagej.init('sc.fiji:fiji', headless=False)

        self.ui.thumbnailWidget.setViewMode(QListWidget.IconMode)
        self.ui.thumbnailWidget.setIconSize(QSize(100, 100))
        self.ui.thumbnailWidget.setResizeMode(QListWidget.Adjust)
        self.ui.thumbnailWidget.setSpacing(10)
        self.ui.thumbnailWidget.itemClicked.connect(self.on_image_clicked)
        # Ensure the layout object is correctly assigned from Qt Designer
        layout = QVBoxLayout()
        self.ui.centralwidget_2.setLayout(layout)

       # Create the ImageJ container widget
        self.imagejContainer = QWidget(self)

       # Set the QSizePolicy for the imagejContainer to Fixed
        size_policy = QSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.imagejContainer.setSizePolicy(size_policy)  # Fix size of the widget

       # Add the imagejContainer to the layout
        layout.addWidget(self.imagejContainer)


        # Create an instance of ImageGraphicsView without adding it to a layout
        self.image_container = ImageGraphicsView(self)

        # Set the geometry to match the original image container's size and position
        self.image_container.setGeometry(self.ui.imageContainer.geometry())

        # Remove the original imageContainer widget (if it's no longer needed)
        self.ui.imageContainer.setParent(None)

        # Set up a QGraphicsScene
        self.scene = QGraphicsScene(self)
        self.image_container.setScene(self.scene)

        # Initialize ImageUploader
        self.image_uploader = ImageUploader()
        all_images = self.image_uploader.get_all_images()  # Get all saved images
        self.display_thumbnails(all_images)

        # Connect the button to open image dialog and save images
        # self.ui.uploadButton.clicked.connect(self.open_image_dialog)


        # Connect the analyze button to process the image
        self.ui.analyzeButton.clicked.connect(self.process_image)

        # Connect the crop button to enable ellipse drawing

        # Connect the save button to save the cropped image
        self.ui.saveImageButton.clicked.connect(self.save_results)
        self.ui.excelButton.clicked.connect(self.export_results)
        self.ui.deleteColonies.clicked.connect(self.delete_colonies)



        # self.ui.nextButton.clicked.connect(self.next_image)
        self.ui.loadImageButton.clicked.connect(self.open_image_with_rois)
        self.ui.addFolderButton.clicked.connect(self.upload_images)

        self.ui.DeleteButton.clicked.connect(self.delete_image)
        self.ui.deleteAll.clicked.connect(self.delete_all_images)

        self.pixmap = None
        self.file_path = None
        self.ellipse_item = None
        self.image = None
        self.image_item = None
        self.drawing_enabled = False


    def export_results(self):
        summary_path = "final_summary.xlsx"

        if not os.path.exists(summary_path):
            QMessageBox.critical(self, "Error", "No data to export: 'final_summary.xlsx' file not found.")
            return

        df_summary_sheets = pd.read_excel(summary_path, sheet_name=None, engine='openpyxl')

        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(
            title="Select an Excel file",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )

        if not file_path:
            print("No file selected.")
            return

        try:
            existing_sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
            updated_sheets = {}

            for sheet_name, df_summary in df_summary_sheets.items():
                # If sheet exists, merge and replace by colony_label
                if sheet_name in existing_sheets:
                    df_existing = existing_sheets[sheet_name]

                    # Ensure both DataFrames have the same columns
                    all_columns = sorted(set(df_existing.columns).union(df_summary.columns))
                    df_existing = df_existing.reindex(columns=all_columns, fill_value=None)
                    df_summary = df_summary.reindex(columns=all_columns, fill_value=None)

                    # Set colony_label as index to allow direct replacement
                    df_existing.set_index("colony_label", inplace=True, drop=False)
                    df_summary.set_index("colony_label", inplace=True, drop=False)

                    # Update or add rows from df_summary
                    df_existing.update(df_summary)
                    df_combined = pd.concat([df_existing, df_summary[~df_summary.index.isin(df_existing.index)]])
                    df_combined = df_combined.reset_index(drop=True)

                    updated_sheets[sheet_name] = df_combined
                else:
                    updated_sheets[sheet_name] = df_summary

            # Preserve other sheets not touched by the summary
            for sheet_name, df_existing in existing_sheets.items():
                if sheet_name not in updated_sheets:
                    updated_sheets[sheet_name] = df_existing

            # Write everything back
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                for sheet_name, df in updated_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"Summary merged and saved to {file_path} with all sheets updated.")

        except FileNotFoundError:
            print(f"File not found: {file_path}")
        except Exception as e:
            print(f"An error occurred: {e}")




    def delete_colonies(self):

        if not hasattr(self, 'current_image_path') or not self.current_image_path:
            print("No image selected.")
            return
        macro = """
        run("ROI Manager...");
        roiCountBefore = roiManager("count");
        run("Add to Manager");
        roiCountAfter = roiManager("count");

        if (roiCountAfter == roiCountBefore) {
            return;
        } else {
            roiManager("Deselect");
            roiManager("Select", roiCountAfter - 1);
            getSelectionCoordinates(xNew, yNew);

            for (i = roiCountAfter - 2; i >= 0; i--) {
                roiManager("Select", i);
                getSelectionCoordinates(xOld, yOld);

                inside = true;
                for (j = 0; j < xOld.length; j++) {
                    if (!pointInPolygon(xOld[j], yOld[j], xNew, yNew)) {
                        inside = false;
                        break;
                    }
                }

                if (inside) {
                    roiManager("Delete");
                }
            }

            roiCountFinal = roiManager("count");
            roiManager("Select", roiCountFinal - 1);
            roiManager("Delete");
        }

        function pointInPolygon(px, py, polyX, polyY) {
            crossings = 0;
            n = polyX.length;
            for (i = 0; i < n; i++) {
                j = (i + 1) % n;
                if (((polyY[i] > py) != (polyY[j] > py)) &&
                    (px < (polyX[j] - polyX[i]) * (py - polyY[i]) / (polyY[j] - polyY[i]) + polyX[i])) {
                    crossings++;
                }
            }
            return (crossings % 2 == 1);
        }
        """
        self.ij.py.run_macro(macro)



    def delete_all_images(self):
        """Delete all images and their associated ROI and CSV data."""
        reply = QMessageBox.question(
            self,
            "Delete All Data",
            "Are you sure you want to delete all images and associated ROI and CSV data?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Define directories
                app_images_dir = os.path.join(os.getcwd(), 'app_images')
                app_ROI_dir = os.path.join(os.getcwd(), 'app_ROI')
                app_csv_dir = os.path.join(os.getcwd(), 'app_csv')

                # Delete contents of each directory
                for dir_path in [app_images_dir, app_ROI_dir, app_csv_dir]:
                    if os.path.exists(dir_path):
                        for filename in os.listdir(dir_path):
                            file_path = os.path.join(dir_path, filename)
                            try:
                                if os.path.isfile(file_path) or os.path.islink(file_path):
                                    os.unlink(file_path)
                                elif os.path.isdir(file_path):
                                    shutil.rmtree(file_path)
                            except Exception as e:
                                raise Exception(f"Failed to delete {file_path}: {e}")

                # Delete the final summary Excel file
                final_summary_path = os.path.join(os.getcwd(), "final_summary.xlsx")
                if os.path.exists(final_summary_path):
                    try:
                        os.remove(final_summary_path)
                        print("Final summary Excel file deleted.")
                    except Exception as e:
                        raise Exception(f"Failed to delete final summary file: {e}")

                QMessageBox.information(self, "Success", "All images and associated data have been deleted.")

                self.image_uploader.saved_hashes = {}

                # Delete metadata JSON file
                metadata_path = os.path.join("metadata", "saved_hashes.json")
                if os.path.exists(metadata_path):
                    try:
                        os.remove(metadata_path)
                        print("Metadata JSON file deleted.")
                    except Exception as e:
                        raise Exception(f"Failed to delete metadata JSON file: {e}")

                self.image_uploader.load_images_into_widget()
                all_images = self.image_uploader.get_all_images()
                self.display_thumbnails(all_images)
                # Clear current image display
                self.scene.clear()
                self.current_image_path = None
                self.pixmap = None

                QMessageBox.information(self, "Deleted", "Image and associated ROI (if any) deleted successfully.")

                if self.hwnd:
                    print(self.hwnd)
                    for elem in self.hwnd:
                        win32gui.PostMessage(elem, win32con.WM_CLOSE, 0, 0)
                    self.hwnd = []
                colony_label = ""
                project_id = ""
                self.ui.colonyIdInput.setText(colony_label)
                self.ui.projectIdInput.setText(project_id)

            except Exception as e:
                QMessageBox.critical(self, "Error", f"An error occurred while deleting: {e}")




    def delete_image(self):
        """Delete the selected image and its associated ROI file."""
        reply = QMessageBox.question(
            self,
            "Delete Image",
            "Are you sure you want to delete this image and its associated ROI data?",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                # Delete the image file
                if os.path.exists(self.current_image_path):
                    os.remove(self.current_image_path)
                    print(f"Deleted image: {self.current_image_path}")
                else:
                    print("Image file does not exist.")

                # Try to delete the associated ROI file
                base_name = os.path.splitext(os.path.basename(self.current_image_path))[0]
                roi_file_path = os.path.join("app_ROI", f"{base_name}_ROI.zip")

                if os.path.exists(roi_file_path):
                    os.remove(roi_file_path)
                    print(f"Deleted ROI file: {roi_file_path}")
                else:
                    print("Associated ROI file not found.")

                # Remove the deleted image from saved_hashes dictionary in memory
                keys_to_remove = []
                for key, value in self.image_uploader.saved_hashes.items():
                    if value["path"] == self.current_image_path:
                        keys_to_remove.append(key)
                for key in keys_to_remove:
                    del self.image_uploader.saved_hashes[key]
                    print(f"Removed {self.current_image_path} and its metadata from saved_hashes.")

                # Also remove the JSON metadata entry from the JSON file
                metadata_path = os.path.join("metadata", "saved_hashes.json")
                if os.path.exists(metadata_path):
                    import json
                    with open(metadata_path, "r") as f:
                        metadata = json.load(f)

                    # Remove keys matching current image path
                    keys_to_remove_from_file = []
                    for key, value in metadata.items():
                        if value.get("path") == self.current_image_path:
                            keys_to_remove_from_file.append(key)
                    for key in keys_to_remove_from_file:
                        metadata.pop(key, None)

                    # Write updated metadata back to the JSON file
                    with open(metadata_path, "w") as f:
                        json.dump(metadata, f, indent=4)
                    print(f"Removed metadata JSON entry for {self.current_image_path}")

                # Refresh the thumbnail list (this also ensures the list widget updates)
                self.image_uploader.load_images_into_widget()
                all_images = self.image_uploader.get_all_images()
                self.display_thumbnails(all_images)

                # Clear current image display
                self.scene.clear()
                self.current_image_path = None
                self.pixmap = None

                if self.hwnd:
                    print(self.hwnd)
                    for elem in self.hwnd:
                        win32gui.PostMessage(elem, win32con.WM_CLOSE, 0, 0)
                    self.hwnd = []

                colony_label = ""
                project_id = ""
                self.ui.colonyIdInput.setText(colony_label)
                self.ui.projectIdInput.setText(project_id)

                QMessageBox.information(self, "Deleted", "Image and associated ROI (if any) deleted successfully.")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"An error occurred while deleting: {e}")



    def on_image_clicked(self, item):


        self.ij.py.run_macro("""
        run("Clear Results", "");
        roiManager("Reset");""")
        """Handles click events on the list item."""
        image_path = item.data(Qt.UserRole)  # Retrieve the stored image path
        print(f"Image clicked: {image_path}")  # Debug: Check which image was clicked

        if self.hwnd:
            print(self.hwnd)
            for elem in self.hwnd:
                    win32gui.PostMessage(elem, win32con.WM_CLOSE, 0, 0)
            self.hwnd = []

        # Running the macro via ImageJ Python API

        # Running the macro via ImageJ Python API


        # Set the current image path
        self.current_image_path = image_path

        # Display the image in the PyQt GUI
        self.display_image(image_path)

        colony_label = ""
        project_id = ""


        for meta in self.image_uploader.saved_hashes.values():
            if meta.get("path") == self.current_image_path:
                colony_label = meta.get("colony_label", "")
                project_id = meta.get("project_id", "")
                break

        self.ui.colonyIdInput.setText(colony_label)
        self.ui.projectIdInput.setText(project_id)


    def display_image(self, image_path):
       """Displays the clicked image in the main view."""
       image = Image.open(image_path)

       # Convert the image to QPixmap to display in the QGraphicsView
       image = image.convert("RGB")  # Ensure it's in RGB mode
       data = image.tobytes("raw", "RGB")  # Convert to raw byte data
       qim = QImage(data, image.width, image.height, image.width * 3, QImage.Format_RGB888)

       # Convert QImage to QPixmap
       pixmap = QPixmap.fromImage(qim)

       if pixmap.isNull():
           print("Failed to load image!")  # Debug: Image not loaded properly
           return

       # Clear any existing items in the scene
       self.scene.clear()

       # Create a QGraphicsPixmapItem with the selected image and add it to the scene
       self.image_item = QGraphicsPixmapItem(pixmap)
       self.scene.addItem(self.image_item)

       # Fit the image to the QGraphicsView initially
       self.image_container.fitInView(self.image_item, Qt.KeepAspectRatio)

       # Store the pixmap for potential future use
       self.pixmap = pixmap





    def upload_images(self):
        saved_images = self.image_uploader.upload_images()
        if saved_images:
            print(f"Saved images: {saved_images}")
            # Use get_all_images to display both newly saved and already existing images
            all_images = self.image_uploader.get_all_images()  # Get all images (including previously saved)
            self.display_thumbnails(all_images)



    def save_results(self):
        if not hasattr(self, 'current_image_path') or not self.current_image_path:
              print("No image selected.")
              return

        self.add_colonies()  # Your function to prepare colonies

          # --- Save ROI ---
        image_name = os.path.splitext(os.path.basename(self.current_image_path))[0]
        roi_filename = f"{image_name}_ROI.zip"
        app_dir = os.path.dirname(os.path.abspath(self.current_image_path))
        base_dir = os.path.dirname(app_dir)
        roi_dir = os.path.join(base_dir, "app_ROI")
        os.makedirs(roi_dir, exist_ok=True)
        file_path = os.path.join(roi_dir, roi_filename)
        escaped_path = file_path.replace("\\", "\\\\")  # Escape for macro

        macro_roi = f'roiManager("Save", "{escaped_path}");'
        self.ij.py.run_macro(macro_roi)
        print(f"ROI saved to {file_path}")

        # Metadata extraction
        project_id = "UNSET"
        colony_label = "Unknown"
        for meta in self.image_uploader.saved_hashes.values():
            if meta.get("path") == self.current_image_path:
                project_id = meta.get("project_id") or "UNSET"
                colony_label = meta.get("colony_label") or "Unknown"
                break

        # Run macro to save Results window as CSV
        os.makedirs("app_csv", exist_ok=True)
        self.ij.py.run_macro("""
            if (isOpen("Results")) {
                saveAs("Results", "app_csv/summary_temp.csv");
            }
        """)

        results_csv = "app_csv/summary_temp.csv"

        # Check if CSV exists and read it
        if not os.path.exists(results_csv):
            print(f"{results_csv} not found. No results saved.")
            return

        try:
            df = pd.read_csv(results_csv)
        except Exception as e:
            print(f"Failed to read CSV {results_csv}: {e}")
            return

        # Calculate summary stats assuming 'Area' column in Results
        if df.empty or "Area" not in df.columns:
            number_of_colonies = 0
            average_size = 0.0
            std_dev_size = 0.0
        else:
            number_of_colonies = len(df)
            average_size = round(df["Area"].mean(), 3)
            std_dev_size = round(df["Area"].std(ddof=1), 3) if number_of_colonies > 1 else 0.0

        final_row = {
            "project_id": project_id,
            "colony_label": colony_label,
            "image_file_name": os.path.basename(self.current_image_path),
            "number_of_colonies": number_of_colonies,
            "average_size": average_size,
            "std_dev_size": std_dev_size
        }

        # Save to Excel (replace sheet data if sheet exists)
        excel_path = "final_summary.xlsx"
        sheet_name = project_id[:31]  # max sheet name length

        if os.path.exists(excel_path):
            book = load_workbook(excel_path)
        else:
            book = Workbook()
            if 'Sheet' in book.sheetnames:
                std = book['Sheet']
                book.remove(std)

        if sheet_name in book.sheetnames:
            ws = book[sheet_name]

            # Read existing sheet into DataFrame
            data = ws.values
            columns = next(data)  # header row
            df_existing = pd.DataFrame(data, columns=columns)

            # Check if colony_label already exists in that sheet
            if colony_label in df_existing['colony_label'].values:
                # Update the existing row for this colony_label
                df_existing.loc[df_existing['colony_label'] == colony_label, :] = list(final_row.values())
            else:
                # Append new row
                df_existing = pd.concat([df_existing, pd.DataFrame([final_row])], ignore_index=True)

            # Clear the worksheet and rewrite with updated df
            ws.delete_rows(1, ws.max_row)
            ws.append(list(df_existing.columns))
            for row in df_existing.itertuples(index=False):
                ws.append(row)

        else:
            ws = book.create_sheet(title=sheet_name)
            ws.append(list(final_row.keys()))  # Write header
            ws.append(list(final_row.values()))

        book.save(excel_path)
        print(f"Results saved to Excel sheet '{sheet_name}' in {excel_path}")



    def open_image_with_rois(self):

       if self.hwnd:
           QMessageBox.information(
               self,
               "Image Already Loaded",
               "An image is already loaded. Please close it before opening another."
           )
           return
       """Auto-load the ROI set associated with the current image."""
       if not hasattr(self, 'current_image_path') or not self.current_image_path:
           print("No image selected.")
           return

       image_path = self.current_image_path.replace("\\", "/")
       image_name = os.path.basename(image_path)
       image_stem = os.path.splitext(image_name)[0]
       roi_dir = os.path.join(os.path.dirname(image_path), "../app_ROI")
       roi_dir = os.path.abspath(roi_dir)
       roi_path = os.path.join(roi_dir, f"{image_stem}_ROI.zip").replace("\\", "/")

       if not os.path.exists(image_path):
           print("Image file does not exist.")
           return

       if not os.path.exists(roi_path):
           print(f"ROI file not found: {roi_path}")
           return

       # Load image as array with OpenCV
       qimage = self.pixmap.toImage()

       # Convert QImage to numpy array
       width = qimage.width()
       height = qimage.height()
       channels = 4  # Assuming RGBA format
       ptr = qimage.bits()
       arr = np.array(ptr).reshape(height, width, channels)
       # image = arr

       # Run preprocessing only
       cropped = self.preprocess_image(arr)
       if cropped is None:
           print("Preprocessing failed.")
           return

       # Convert to ImageJ image and show
       imp = self.ij.py.to_java(cropped)
       self.ij.ui().show(imp)
       hwnd = self.find_imagej_hwnd()
       if hwnd:
           print(f"Found ImageJ HWND: {hwnd}")
           self.hwnd.append(hwnd)
           try:
               win = QWindow.fromWinId(hwnd)  # Convert hwnd to QWindow
               container = self.createWindowContainer(win)  # Create a container for QWindow

               # Parent the container to centralwidget_2
               self.container = container
               self.container.setParent(self.ui.centralwidget_2)

               # Set the geometry to match the parent widget's size (top-left)
               self.container.setGeometry(self.ui.centralwidget_2.rect())  # Fills the space
               self.container.show()

           except Exception as e:
               print(f"Error embedding ImageJ window: {e}")
       else:
           print("Failed to find ImageJ window")

       self.ij.py.run_macro("setTool('freehand');",)
       # Open ROIs
       macro = f"""
       run("ROI Manager...");
       roiManager("Reset");
       roiManager("Open", "{roi_path}");
       //roiManager("Show All");
       roiManager("Show All with labels");
       roiManager("Deselect");
       roiManager("Measure");

       """
       self.ij.py.run_macro(macro)
       # self.add_colonies()








    def add_colonies(self):


        macro = """
        //run("Clear Results", "");
        roiCount = roiManager("count");

        run("Add to Manager");


        if (!isOpen("ROI Manager")) {
            run("ROI Manager...");
        }

        roiCount = roiManager("count");
        roiManager("Select", roiCount - 1);

        // Debug: Print selection type
        isComposite = (selectionType() == 9); // Check for composite selection

        if (isComposite) {
            roiManager("Split");
            wait(200); // Give time for processing
        }

        // Get the updated ROI count
        roiCountAfter = roiManager("count");

        if (!isComposite) {
            // Process single ROI directly
            roiManager("Select", roiCount - 1);
            run("Add Selection...");
        } else {
            // Remove original composite ROI
            roiManager("Select", roiCount - 1);
            roiManager("Delete");
            }
        roiCountAfter = roiManager("count"); // Update count


        run("Clear Results", "");
        roiManager("Deselect");
        roiManager("Measure");
        // Process each split ROI
        //for (i = roiCountAfter - 1; i >= 0; i--) {
            //roiManager("Select", i);
            //roiManager("Measure");

       // }
        //for (i = roiCountAfter - 1; i >= roiCount-1; i--) {
           // roiManager("Select", i);
           // run("Add Selection...");
      // }
       // roiManager("Deselect");
       // run("Select None");
    """

        self.ij.py.run_macro(macro)

    def find_imagej_hwnd(self, timeout=5.0):
        start_time = time.time()
        while time.time() - start_time < timeout:
            def enum_handler(hwnd, result):
                title = win32gui.GetWindowText(hwnd)
                try:
                    print(f"Window title: {title}")  # For debugging
                except UnicodeEncodeError:
                    print(f"Window title (unicode error): {title.encode('ascii', 'ignore').decode('ascii')}")
                if "(V)" in title:
                    result.append(hwnd)

            hwnds = []
            win32gui.EnumWindows(enum_handler, hwnds)
            if hwnds:
                return hwnds[0]

        return None









    def preprocess_image(self, arr):


        if arr.shape[2] == 4:
            arr = arr[:, :, :3]

        resize_factor = 0.1
        resized_image = cv2.resize(arr, (0, 0), fx=resize_factor, fy=resize_factor)
        gray_resized = cv2.cvtColor(resized_image, cv2.COLOR_BGR2GRAY)

        circles_resized = cv2.HoughCircles(
            gray_resized, cv2.HOUGH_GRADIENT,
            dp=1.3, minDist=100, param1=50, param2=30,
            minRadius=30, maxRadius=100
        )

        if circles_resized is not None:
            circles_resized = np.round(circles_resized[0, :]).astype("int")
            image_center = (arr.shape[1] // 2, arr.shape[0] // 2)
            most_centered_circle = min(
                circles_resized,
                key=lambda c: np.sqrt((int(c[0]/resize_factor) - image_center[0])**2 + (int(c[1]/resize_factor) - image_center[1])**2)
            )
            x, y, r = (int(most_centered_circle[0] / resize_factor),
                       int(most_centered_circle[1] / resize_factor),
                       int(most_centered_circle[2] / resize_factor))

            mask = np.zeros_like(arr)
            cv2.circle(mask, (x, y), r, (255, 255, 255), -1)
            result = cv2.bitwise_and(arr, mask)

            x1, y1 = max(0, x - r), max(0, y - r)
            x2, y2 = min(arr.shape[1], x + r), min(arr.shape[0], y + r)
            cropped_image = result[y1:y2, x1:x2]

            cropped_image = cropped_image.dot([0.299, 0.587, 0.114])
            cropped_image = np.clip(cropped_image, 0, 255).astype(np.uint8)


            # downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            # output_path = os.path.join(downloads_dir, "cropped_image.png")
            # cv2.imwrite(output_path, cropped_image)
            # print(f"Saved cropped image to: {output_path}")

            return cropped_image
        else:
            print("No circles found.")
            return None






    def process_image(self):


        if self.hwnd:
            QMessageBox.information(
                self,
                "Image Already Loaded",
                "An image is already loaded. Please close it before opening another."
            )
            return
        if self.pixmap is None:
            QMessageBox.information(None, "Error", "No image loaded")
            return

        project_id = self.ui.projectIdInput.text().strip()
        colony_label = self.ui.colonyIdInput.text().strip()

        if not project_id or not colony_label:
            QMessageBox.information(
                self,
                "Invalid Project ID or Colony Label",
                "Enter a unique Project ID and Colony Label."
            )
            return


        # Check for duplicate colony label within the same project_id
        for meta in self.image_uploader.saved_hashes.values():
            # meta is a dict with keys "path", "project_id", "colony_label"
            if meta["project_id"] == project_id and meta["colony_label"] == colony_label and self.current_image_path != meta["path"]:
                existing_path = meta["path"]
                QMessageBox.information(
                    self,
                    "Duplicate Colony Label",
                    f"Colony Label '{colony_label}' already exists for Project ID '{project_id}'.\n"
                    f"Existing image path:\n{existing_path}"
                )
                return  # Stop processing due to duplicate




        image_hash = self.image_uploader._get_image_hash(self.current_image_path)

        self.image_uploader.saved_hashes[image_hash]["project_id"] = project_id
        self.image_uploader.saved_hashes[image_hash]["colony_label"] = colony_label


        try:
            with open("metadata/saved_hashes.json", "w") as f:
                json.dump(self.image_uploader.saved_hashes, f, indent=4)
        except Exception as e:
            QMessageBox.warning(self, "Save Failed", f"Could not save metadata: {e}")

        # Convert QPixmap to QImage
        qimage = self.pixmap.toImage()

        # Convert QImage to numpy array
        width = qimage.width()
        height = qimage.height()
        channels = 4  # Assuming RGBA format
        ptr = qimage.bits()
        arr = np.array(ptr).reshape(height, width, channels)
        # image = arr
        # if arr.shape[2] == 4:
        #     arr = arr[:, :, :3].dot([0.299, 0.587, 0.114])  # Convert to grayscale
        # arr = np.clip(arr, 0, 255).astype(np.uint8)



        # Check if the image is loaded properly
        if arr is None:
            print("Error: The image could not be loaded.")
        else:
             #Resize the image to reduce computation
            resize_factor = 0.1  # Resize to 10% of the original size (you can adjust this)
            resized_image = cv2.resize(arr, (0, 0), fx=resize_factor, fy=resize_factor)

            # Convert resized image to grayscale
            gray_resized = cv2.cvtColor(resized_image, cv2.COLOR_BGR2GRAY)


            circles_resized = cv2.HoughCircles(
                gray_resized,
                cv2.HOUGH_GRADIENT,
                dp=1.3,
                minDist=100,
                param1=50,
                param2=30,
                minRadius=30,
                maxRadius=100
            )

      # If circles are detected, find the most centered circle
            if circles_resized is not None:
                circles_resized = np.round(circles_resized[0, :]).astype("int")

                # Calculate the center of the original image
                height, width = arr.shape[:2]
                image_center = (width // 2, height // 2)

                # Variable to store the most centered circle
                most_centered_circle = None
                min_distance = float('inf')

                # Find the circle closest to the center of the image
                for (x_resized, y_resized, r_resized) in circles_resized:
                    # Scale the circle parameters back to the original image size
                    x_original = int(x_resized / resize_factor)
                    y_original = int(y_resized / resize_factor)
                    r_original = int(r_resized / resize_factor)

                    # Calculate the Euclidean distance from the center of the image to the circle center
                    distance = np.sqrt((x_original - image_center[0]) ** 2 + (y_original - image_center[1]) ** 2)

                    # If this circle is the most centered so far, store it
                    if distance < min_distance:
                        most_centered_circle = (x_original, y_original, r_original)
                        min_distance = distance

                # If we found a most centered circle, crop the image inside the circle
                if most_centered_circle:
                    x, y, r = most_centered_circle


                    # Step 5: Create a mask for the circle
                    mask = np.zeros_like(arr)
                    cv2.circle(mask, (x, y), r, (255, 255, 255), -1)  # White circle on black background

                    # Step 6: Apply the mask to the original image
                    result = cv2.bitwise_and(arr, mask)

                    # Step 7: Crop the region inside the circle
                    # Create a bounding box around the circle
                    x1 = max(0, x - r)
                    y1 = max(0, y - r)
                    x2 = min(arr.shape[1], x + r)
                    y2 = min(arr.shape[0], y + r)

                    cropped_image = result[y1:y2, x1:x2]

                    # Step 8: Display the cropped result

                else:
                    print("No circles detected.")
            else:
                print("No circles detected.")





        if cropped_image.shape[2] == 4:  # Check if the image has 4 channels (RGBA)
            # Discard the alpha channel (use only RGB channels)
            cropped_image = cropped_image[:, :, :3]

        # Convert to grayscale (if RGB image)
        cropped_image = cropped_image.dot([0.299, 0.587, 0.114])  # RGB to grayscale conversion
        cropped_image = np.clip(cropped_image, 0, 255).astype(np.uint8)  # Ensure values are within 0-255 range

        # Convert numpy array to ImagePlus object
        imp = self.ij.py.to_java(cropped_image)

        # Show the image in ImageJ


        # Running the macro via ImageJ Python API
        # self.ij.py.run_macro(macro)


        # # Running the macro via ImageJ Python API
        # self.ij.py.run_macro(macro)
        self.ij.ui().show(imp)

        # time.sleep(1)  # Give time to draw



        # Handle the resizing event to reset position and avoid jumping around


        self.ij.py.run_macro('roiManager("Reset");')

        self.ij.py.run_macro("run('8-bit');",)  # Convert to 8-bit

        self.ij.py.run_macro("run('Convert to Mask');",)
        self.ij.py.run_macro('setThreshold(0, 0);')
        self.ij.py.run_macro("run('Convert to Mask');",)

        self.ij.py.run_macro("run('Remove Outliers...', 'radius=2 threshold=50 which=Bright');",)
        for i in range(2):
            self.ij.py.run_macro("run('Despeckle');",)

        self.ij.py.run_macro("run('Watershed');",)
        self.ij.py.run_macro("run('Analyze Particles...', 'size=70-50000 circularity=0.70-1.00 display exclude summarize overlay add');",)

        # project_id = self.ui.projectIdInput.text().strip()
        # colony_label = self.ui.colonyIdInput.text().strip()

        # if not project_id:
        #     project_id = "UNSET"
        # if not colony_label:
        #     colony_label = "Unknown"


        # self.ij.py.run_macro("""
        #     // Save Summary table if open
        #     if (isOpen("Summary")) {
        #         selectWindow("Summary");
        #         saveAs("Results", "summary_temp.csv");
        #     }
        #     // Save Results table if open
        #     if (isOpen("Results")) {
        #         selectWindow("Results");
        #         saveAs("Results", "results_table.csv");
        #     }
        # """)
        # df_results = pd.read_csv("results_table.csv")
        # sizes = df_results["Area"]
        # std_dev_size = round(df_results["Area"].std(ddof=1), 3)


        # # Load the CSV into pandas
        # df_summary = pd.read_csv("summary_temp.csv")
        # # copy summary results from csv to excel


        # image_file_name = os.path.basename(self.current_image_path)


        # try:
        #     number_of_colonies = int(df_summary.at[0, 'Count'])
        #     average_size = float(df_summary.at[0, 'Average Size'])
        # except KeyError:
        #     print("Expected columns missing.")
        #     return


        # final_row = pd.DataFrame([{
        #     "project_id": project_id,
        #     "image_file_name": image_file_name,
        #     "colony_label": colony_label,
        #     "number_of_colonies": number_of_colonies,
        #     "average_size": average_size,
        #     "std_dev_size": std_dev_size
        # }])

        # # Save to final_summary.csv (append if exists)
        # csv_path = "final_summary.csv"
        # if os.path.exists(csv_path):
        #     final_row.to_csv(csv_path, mode='a', header=False, index=False)
        # else:
        #     final_row.to_csv(csv_path, index=False)

        self.ij.py.run_macro("setTool('freehand');",)
        self.ij.ui().show(imp)



        hwnd = self.find_imagej_hwnd()
        if hwnd:
            print(f"Found ImageJ HWND: {hwnd}")
            self.hwnd.append(hwnd)
            try:
                win = QWindow.fromWinId(hwnd)  # Convert hwnd to QWindow
                container = self.createWindowContainer(win)  # Create a container for QWindow

                # Parent the container to centralwidget_2
                self.container = container
                self.container.setParent(self.ui.centralwidget_2)

                # Set the geometry to match the parent widget's size (top-left)
                self.container.setGeometry(self.ui.centralwidget_2.rect())  # Fills the space
                self.container.show()

            except Exception as e:
                print(f"Error embedding ImageJ window: {e}")
        else:
            print("Failed to find ImageJ window")

        self.ij.py.run_macro("""
        if (!isOpen("ROI Manager")) {
            run("ROI Manager...");
        }
        roiCount = roiManager("count");
        if (roiCount > 0) {
            roiManager("Select", 0);
            roiManager("Update");
        }
        """)



    def wheelEvent(self, event):
        if hasattr(self, "container"):
            self.container.move(0, 0)  # Move to top-left corner explicitly if needed
            # Get the current mouse position within the container


    def resizeEvent(self, event):
         """
         Ensure that the ImageJ window stays centered (top-left position)
         """
         super().resizeEvent(event)

         # Ensure the container (ImageJ window) stays fixed in the top-left of the parent widget
         if hasattr(self, "container"):
             parent_rect = self.ui.centralwidget_2.rect()

             # Align the container to the top-left corner of the parent widget
             self.container.setGeometry(parent_rect)  # Align with the parent widget's geometry
             self.container.move(0, 0)  # Move to top-left corner explicitly if needed
             print("Container resized and moved to top-left.")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    widget = MainWindow()
    widget.show()
    sys.exit(app.exec())
